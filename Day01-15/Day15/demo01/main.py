import datetime

from openpyxl import Workbook
#Python的openpyxl模块让我们可以在Python程序中读取和修改Excel电子表格，由于微软从Office 2007开始使用了新的文件格式，
# 这使得Office Excel和LibreOffice Calc、OpenOffice Calc是完全兼容的，这就意味着openpyxl模块也能处理来自这些软件生成的电子表格。
#pip install openpyxl

wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")


# penpyxl是一个用于读取和编写Excel 2010 xlsx/xlsm/xltx/xltm文件的库。
# https://openpyxl.readthedocs.io/en/stable/
# 1、openpyxl支持.xlsx文件的读写。
# 2、支持Excel操作。
# 3、加载大.xlsx文件可以使用read_only模式。
# 4、写入大.xlsx文件可以使用write_only模式。

